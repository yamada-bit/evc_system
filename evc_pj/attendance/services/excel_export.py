"""
勤務表 Excel 出力サービス。

attendance DB のデータを元に勤務表テンプレートへ書き込み、
BytesIO として返す。Google Calendar 依存を完全に排除した版。

【テンプレートファイルの配置】
  settings.ATTENDANCE_EXCEL_TEMPLATE_DIR/勤務表テンプレート_{year}.xlsx
  例: /data_root/data/excel_template/勤務表テンプレート_2026.xlsx
  ※ 1〜3月は前年度テンプレートを使用（4月始まりの年度区分）

【Excelシート構成（テンプレート依存）】
  Sheet1 (報告): I3=年, I4=月, B5=氏名, 9行目〜 が日次データ
  Sheet2 (精算): K3=年, K4=月, 9行目〜 が日次データ

  各日次行 (報告シート):
    B列 = 出勤時刻  (time)
    C列 = 退勤時刻  (time)
    E列 = 勤務場所  (str)
    G列 = 業務内容  (str)
"""
import datetime
import logging
import os
import platform
import re
from io import BytesIO

import openpyxl
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill

logger = logging.getLogger(__name__)

# Excel でフォーミュラとして解釈されるプレフィックス文字（csv_safe() と統一）
_FORMULA_PREFIX_CHARS = frozenset('=+-@\t\r')


def _sanitize_cell_value(value: str) -> str:
    """=, +, -, @, タブ, 復帰文字で始まる文字列をフォーミュラとして実行されないよう先頭に ' を付ける。"""
    if value and value[0] in _FORMULA_PREFIX_CHARS:
        return "'" + value
    return value


if platform.system() == 'Windows':
    _DEFAULT_TEMPLATE_DIR = '../data/excel_template'
else:
    _DEFAULT_TEMPLATE_DIR = '/data_root/data/excel_template'


def _get_template_dir() -> str:
    return getattr(settings, 'ATTENDANCE_EXCEL_TEMPLATE_DIR', _DEFAULT_TEMPLATE_DIR)


# ---------------------------------------------------------------------------
# attendance DB → Excel 向けデータ変換
# ---------------------------------------------------------------------------

def build_attendance_for_excel(user_id: str, year: int, month: int) -> dict:
    """
    attendance DB から指定ユーザー・年月の勤怠データを取得し、
    update_attendance_to_excel() が期待する形式の辞書に変換する。

    戻り値:
      { date(year, month, day): {
          'start':    datetime | None,  # 出勤日時 (ローカル時刻)
          'end':      datetime | None,  # 退勤日時 (ローカル時刻)
          'location': str,              # 日報の外出先・勤務場所詳細（Excel E列）
          'title':    str,              # 日報の業務内容（未記入時は空文字）
          'allday':   str,              # '終日' (休暇系) or ''
        }, ...
      }
    """
    # インポートをここで行うことで循環参照を回避する
    from ..models import Attendance, DailyReport
    from ..views._base import using_db

    _LEAVE_TYPES = frozenset({'PAID_LEAVE', 'AM_LEAVE', 'PM_LEAVE', 'COMP_LEAVE'})

    attendances = (
        Attendance.objects.using(using_db)
        .filter(user_id=user_id, work_date__year=year, work_date__month=month)
        .only('work_date', 'work_type', 'clock_in', 'clock_out')
    )

    reports = (
        DailyReport.objects.using(using_db)
        .filter(user_id=user_id, report_date__year=year, report_date__month=month)
        .only('report_date', 'task_summary', 'location_detail')
    )
    # date → (task_summary, location_detail)
    report_map = {
        r.report_date: (r.task_summary or '', r.location_detail or '')
        for r in reports
    }

    result = {}
    for att in attendances:
        is_leave = att.work_type in _LEAVE_TYPES
        task_summary, location_detail = report_map.get(att.work_date, ('', ''))
        result[att.work_date] = {
            'start':    timezone.localtime(att.clock_in)  if att.clock_in  else None,
            'end':      timezone.localtime(att.clock_out) if att.clock_out else None,
            'location': location_detail,   # 日報の「外出先・勤務場所詳細」→ Excel E列
            'title':    task_summary,      # 日報の「業務内容」→ Excel G列
            'allday':   '終日' if is_leave else '',
        }
    return result


# ---------------------------------------------------------------------------
# Excel 書き込みロジック（Kms_Calendar/excel_calendar.py から移植・変更なし）
# ---------------------------------------------------------------------------

def update_attendance_to_excel(attendance: dict, year: int, month: int, user_name: str) -> BytesIO | None:
    """
    テンプレートを読み込み、attendance データを書き込んで BytesIO を返す。
    テンプレートが見つからない場合は None を返す。
    その他の予期しないエラーは呼び出し元に伝播する。
    """
    template_dir = _get_template_dir()
    template_year = year - 1 if month < 4 else year
    template_path = os.path.join(
        template_dir, f'勤務表テンプレート_{template_year}.xlsx'
    ).replace(os.sep, '/')

    if not os.path.exists(template_path):
        logger.warning(f'勤務表テンプレートが見つかりません: {template_path}')
        return None

    wb = openpyxl.load_workbook(template_path)
    sheet1 = wb.worksheets[0]
    sheet2 = wb.worksheets[1]
    if not sheet1 or not sheet2:
        logger.error(f'テンプレートのシート数が不正です: {template_path}')
        return None
    _update_sheet(sheet1, attendance, year, month, 0, user_name)
    _update_sheet(sheet2, attendance, year, month, 1, user_name)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _update_sheet(sheet, attendance: dict, year: int, month: int, index: int, user_name: str) -> bool:
    start_date = datetime.datetime(year, month, 1).date()
    end_date = start_date + relativedelta(months=+1, day=1, days=-1)
    if index == 0:
        sheet['I3'] = year
        sheet['I4'] = month
        last_col = 10
        sheet.title = f'報告{month}月'
        sheet['B5'] = user_name
    else:
        sheet['K3'] = year
        sheet['K4'] = month
        last_col = 12
        sheet.title = f'清算{month}月'
        _update_formula_sheet(sheet, f'報告{month}月')

    row_offset = 9
    last_row = 40
    num_days = (end_date - start_date).days

    for count_date in range(0, num_days + 1):
        target_date = start_date + datetime.timedelta(count_date)
        if target_date in attendance:
            try:
                event = attendance[target_date]
                if event:
                    if index == 0:
                        if event.get('allday') == '終日':
                            sheet[f'B{count_date + row_offset}'] = ''
                            sheet[f'C{count_date + row_offset}'] = ''
                        else:
                            if event['start']:
                                sheet[f'B{count_date + row_offset}'] = event['start'].time()
                            if event['end']:
                                sheet[f'C{count_date + row_offset}'] = event['end'].time()
                        sheet[f'E{count_date + row_offset}'] = _sanitize_cell_value(event.get('location', ''))
                        sheet[f'G{count_date + row_offset}'] = _sanitize_cell_value(event.get('title', ''))
            except Exception:
                logger.exception(f'_update_sheet exception {target_date}')

    # 残りの空白行をクリア
    for i in range(num_days + row_offset + 1, last_row):
        for col in ('A', 'B', 'C', 'D', 'E', 'G'):
            sheet[f'{col}{i}'] = ''
    return True


def _update_formula_sheet(sheet, sheet_name: str) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = cell.value.replace('報告1月', sheet_name)


def _shift_formula(formula: str, inserted_row: int) -> str:
    def shift_match(match):
        col, row = match.groups()
        if row:
            row_num = int(row)
            if row_num >= inserted_row:
                return f'{col}{row_num + 1}'
            return f'{col}{row_num}'
        return col
    return re.sub(r'([A-Z]+)(\d+)', shift_match, formula)


def _shift_formula_after_insert(sheet, inserted_row: int) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = _shift_formula(cell.value, inserted_row)


def _copy_row_with_style(sheet, src_row: int, tgt_row: int, last_col: int) -> None:
    for col in range(1, last_col + 1):
        cell = sheet.cell(row=src_row, column=col)
        new_cell = sheet.cell(row=tgt_row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith('='):
            new_cell.value = cell.value if col == 1 else _shift_formula(cell.value, src_row)
        if cell.has_style:
            new_cell.font = Font(**cell.font.__dict__)
            new_cell.border = Border(**cell.border.__dict__)
            new_cell.fill = PatternFill(**cell.fill.__dict__)
            new_cell.alignment = Alignment(**cell.alignment.__dict__)
            new_cell.number_format = cell.number_format


def _copy_merged_cells(sheet, src_row: int, tgt_row: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row == src_row:
            sheet.merge_cells(
                start_row=tgt_row,
                start_column=merged_range.min_col,
                end_row=tgt_row,
                end_column=merged_range.max_col,
            )


def _increment_formula_row_numbers(formula: str, insert_row: int) -> str:
    def replace_match(match):
        col, dollar_sign, row = match.groups()
        row_num = int(row)
        if row_num >= insert_row:
            return f'{col}{dollar_sign}{row_num + 1}'
        return match.group(0)
    return re.sub(r'(\$?[A-Z]+)(\$?)(\d+)', replace_match, formula)


def _update_conditional_formatting(sheet, insert_row: int) -> None:
    cf_rules = sheet.conditional_formatting._cf_rules
    updated_rules = {}
    for key, rules in cf_rules.items():
        new_rules = []
        for rule in rules:
            if rule and hasattr(rule, 'formula'):
                rule.formula = [
                    _increment_formula_row_numbers(f, insert_row) for f in rule.formula
                ]
            new_rules.append(rule)
        updated_rules[key] = new_rules
    sheet.conditional_formatting._cf_rules = {}
    for key, rules in updated_rules.items():
        sheet.conditional_formatting._cf_rules[key] = rules


# ---------------------------------------------------------------------------
# 公開 API
# ---------------------------------------------------------------------------

def export_attendance_excel(user_id: str, year: int, month: int, user_name: str) -> BytesIO | None:
    """
    指定ユーザー・年月の勤務表 Excel を生成して BytesIO で返す。
    テンプレートが見つからない場合は None を返す。
    """
    attendance = build_attendance_for_excel(user_id, year, month)
    return update_attendance_to_excel(attendance, year, month, user_name)
