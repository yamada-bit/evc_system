import datetime
import logging
import os
import platform
import re
from io import BytesIO

import openpyxl

# from django.conf import settings
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment, Border, Font, PatternFill

if platform.system() == 'Windows':
    EXCEL_TEMPLATE_FOLDER = '../data/excel_template' # 勤務表を作成するテンプレートファイルを保存するフォルダ
else:
    EXCEL_TEMPLATE_FOLDER = '/data_root/data/excel_template' # 勤務表を作成するテンプレートファイルを保存するフォルダ

logger = logging.getLogger(__name__)

"""勤務表をExcelとして出力"""
def calendar_to_excel(grouped_events):
    wb = openpyxl.Workbook()
    ws = wb.active
    if not ws:
        return None

    ws.title = 'Googleカレンダー'
    # ヘッダーを作成
    ws.append(['日付', 'イベント'])
    # for date, event_list in grouped_events.items():
    #     ws.append([date, '\n'.join(event_list)])  # 1セルに複数イベント（改行区切り）
    # 各行にデータを追加
    row_num = 2  # 1行目はヘッダーなので2行目から
    for date, event_list in grouped_events.items():
        event_text = '\n'.join(event_list)  # 1セル内で改行
        # ws.append([date, event_text])
        ws.cell(row_num, 1).value = date
        ws.cell(row_num, 2).value = event_text
        ws.cell(row_num, 2).alignment = Alignment(wrapText=True)
        # イベント数に応じてセルの高さを変更
        num_lines = len(event_list)  # 改行数（イベントの個数）
        ws.row_dimensions[row_num].height = max(15, num_lines * 15)  # 1行あたり15pt

        row_num += 1

    # 列の幅を調整
    ws.column_dimensions['A'].width = 15  # 日付の列
    ws.column_dimensions['B'].width = 40  # イベントの列

    # メモリに保存（BytesIO を使用）
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # ファイルポインタを先頭に戻す

    return output

"""勤務表テンプレートを読み込み、出退勤情報を更新して保存"""
def update_attendance_to_excel(attendance, year, month, user_name):
    # 年毎にテンプレートの祝日データが変更される
    if month < 4:
        template_path =  os.path.join(EXCEL_TEMPLATE_FOLDER, f'勤務表テンプレート_{year-1}.xlsx').replace(os.sep,'/')
    else:
        template_path =  os.path.join(EXCEL_TEMPLATE_FOLDER, f'勤務表テンプレート_{year}.xlsx').replace(os.sep,'/')
    try:
        wb = openpyxl.load_workbook(template_path)
        # インデックス番号を指定してシート取得（インデックス番号は0から始まる）
        sheet1 = wb.worksheets[0]   # 報告
        sheet2 = wb.worksheets[1]   # 精算
        if not sheet1 or not sheet2:
            return None
        update_sheet(sheet1, attendance, year, month, 0, user_name)
        update_sheet(sheet2, attendance, year, month, 1, user_name)

        # メモリに保存（BytesIO を使用）
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # ファイルポインタを先頭に戻す
    except Exception:
        logger.exception('update sheet exception')
        output = None

    return output
# シートの出退勤情報を更新
def update_sheet(sheet, attendance, year, month, index, user_name):
    target_year = year #ut_get_localtoday().year
    target_month = month #ut_get_localtoday().month-1
    start_date = datetime.datetime(target_year, target_month, 1).date()
    end_date = start_date + relativedelta(months=+1, day=1, days=-1)
    if index == 0:
        # I3:年, I4:月
        sheet['I3'] = year
        sheet['I4'] = month
        last_col = 10
        sheet.title = f'報告{month}月'
        sheet['B5'] = user_name
    else:
        # K3:年, K4:月
        sheet['K3'] = year
        sheet['K4'] = month
        last_col = 12
        sheet.title = f'清算{month}月'
        # 数式のシート名の変更
        update_formula_sheet(sheet, f'報告{month}月')
    start_row = 9
    row_offset = 9  # 行のオフセット（複数イベントで行を追加するため）
    last_row = 40   # 日付の最終行（行挿入で結合セルの情報を最終行にコピーする）
    for count_date in range(0, (end_date - start_date).days + 1):
        # ターゲット日（繰り返しのカレント日付）
        target_date = start_date + datetime.timedelta(count_date)
        if target_date in attendance:
            try:
                i = 0
                event = attendance[target_date]
                if event:                       # 勤務イベントを日ごとに整理
                # for i, event in enumerate(attendance[target_date]): # 勤務時間を個別に集計
                    if 0 < i:   # 同一日付で複数データ
                        # 行を挿入してコピー（以降の行を下にシフト）
                        insert_and_copy_row(sheet, count_date + row_offset, count_date + row_offset + 1, last_row, last_col)
                        row_offset += 1
                        last_row += 1
                    if index == 0:  # 報告シート
                        if event.get('allday') == '終日':
                            sheet[f'B{count_date + row_offset}'] = ''
                            sheet[f'C{count_date + row_offset}'] = ''
                        else:
                            # sheet[f'A{count_date + row_offset}'] = target_date.strftime('%Y-%m-%d')  # 日付
                            # sheet[f'B{count_date + row_offset}'] = event['start'].strftime('%#H:%M:00')  # 出勤時間
                            # sheet[f'C{count_date + row_offset}'] = event['end'].strftime('%#H:%M:00')  # 退勤時間
                            sheet[f'B{count_date + row_offset}'] = event['start'].time()  # 出勤時間
                            sheet[f'C{count_date + row_offset}'] = event['end'].time()  # 退勤時間
                            # work_hours = (event['end'] - event['start']).seconds / 3600
                            # sheet[f'D{count_date + row_offset}'] = round(work_hours, 2)  # 勤務時間
                        sheet[f'E{count_date + row_offset}'] = event['location']
                        sheet[f'G{count_date + row_offset}'] = event['title']
                        # sheet[f'I{count_date + row_offset}'] = event['description']
            except Exception:
                logger.exception(f'update sheet exception {target_date}')
    # 明細の残りの空白行
    for i in range(count_date + row_offset + 1, last_row):
        sheet[f'A{i}'] = ''
        sheet[f'B{i}'] = ''
        sheet[f'C{i}'] = ''
        sheet[f'D{i}'] = ''
        sheet[f'E{i}'] = ''
        sheet[f'G{i}'] = ''

    return True

def update_formula_sheet(sheet, sheet_name):
    """既存の数式のシート名の変更"""
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = cell.value.replace('報告1月', sheet_name)

"""src_row のデータと書式をコピーし、tgt_row に挿入（以降の行を下にシフト）"""
def insert_and_copy_row(sheet, src_row, tgt_row, last_row, last_col):
    """
    insert_rows()ではセルに付与された属性情報（定義された名前、書式、条件付き書式）
    ・結合セル・セル上に配置されたイメージ情報・関数式などが追随しない
    """
    # 1. 行を挿入
    sheet.insert_rows(tgt_row)
    # 数式の相対参照を +1する
    shift_formula_after_insert(sheet, tgt_row)  # 数式を更新
    # 2. 書式込みで挿入行にコピー(指定の列まで)
    copy_row_with_style(sheet, src_row, tgt_row, last_col)
    # 3. 結合セルを最終行 +1にコピー
    # 挿入で最終行が下にシフトされるが結合セルは追随しないため結合セルを最終行+1にコピー
    copy_merged_cells(sheet, src_row, last_row + 1)
    # 4. 条件付き書式の影響を受ける行番号を +1する
    update_conditional_formatting(sheet, tgt_row)

"""`insert_rows()` で行を挿入した後に、既存の数式の相対参照を +1する"""
def shift_formula_after_insert(sheet, inserted_row):
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = shift_formula(cell.value, inserted_row)

"""src_row の数式と書式を tgt_row にコピーする（カスタムフォーマット対応）"""
def copy_row_with_style(sheet, src_row, tgt_row, last_col):
    # 日付以外のデータはコピーしない
    # for col in range(1, sheet.max_column + 1):
    for col in range(1, last_col + 1):
        cell = sheet.cell(row=src_row, column=col)
        new_cell = sheet.cell(row=tgt_row, column=col)

        if isinstance(cell.value, str) and cell.value.startswith("="):
            if col == 1:
                new_cell.value = cell.value  # 日付はそのままコピー
            else:
                # 数式を相対参照でコピー
                new_cell.value = shift_formula(cell.value, src_row)
        # else:
        #     new_cell.value = cell.value  # 数式でなければそのままコピー

        # 書式（フォント、罫線、塗りつぶし、配置）をコピー
        if cell.has_style:
            new_cell.font = Font(**cell.font.__dict__)
            new_cell.border = Border(**cell.border.__dict__)
            new_cell.fill = PatternFill(**cell.fill.__dict__)
            new_cell.alignment = Alignment(**cell.alignment.__dict__)
            new_cell.number_format = cell.number_format  # ユーザー定義書式をコピー

    # 行の高さをコピー
    # sheet.row_dimensions[tgt_row].height = sheet.row_dimensions[src_row].height

"""数式のセル参照を相対的に変換する（行のシフト）"""
def shift_formula(formula, inserted_row):
    # 数式のセル参照を +1する（挿入された行より下の行番号を +1）
    if not formula or not formula.startswith("="):
        return formula  # 数式でなければそのまま返す
    def shift_match(match):
        col, row = match.groups()
        if row:  # 行番号がある場合のみ変換
            row = int(row)
            if row >= inserted_row:  # 挿入行より下の参照を+1
                return f"{col}{row+1}"
            return f"{col}{row}"
        return col  # 例: "SUM(A:A)" のような場合は変更しない

    return re.sub(r"([A-Z]+)(\d+)", shift_match, formula)

"""結合セルをコピー（同じ範囲を tgt_row に適用）"""
def copy_merged_cells(sheet, src_row, tgt_row):
    for merged_range in list(sheet.merged_cells.ranges):  # list() でコピーを取る
        if merged_range.min_row == src_row:
            sheet.merge_cells(
                start_row=tgt_row,
                start_column=merged_range.min_col,
                end_row=tgt_row,
                end_column=merged_range.max_col
            )

"""条件付き書式の影響を受ける行番号のみ +1する"""
def update_conditional_formatting(sheet, insert_row):
    # 条件付き書式のルールを取得
    cf_rules = sheet.conditional_formatting._cf_rules

    # 更新後のルールを格納する辞書
    updated_rules = {}

    # 条件付き書式のルールを処理
    for key, rules in cf_rules.items():
        new_rules = []
        for rule in rules:
            if rule:
                # if isinstance(rule, FormulaRule):   # エラーになる
                if hasattr(rule, "formula"):  # "formula" を持つ場合のみ処理
                    updated_formula = []
                    for formula in rule.formula:
                        # 影響を受ける行の行番号のみ+1
                        updated_formula.append(increment_formula_row_numbers(formula, insert_row))
                    rule.formula = updated_formula  # 更新
            new_rules.append(rule)

        updated_rules[key] = new_rules  # ルールを保存

    # 既存の条件付き書式を削除
    sheet.conditional_formatting._cf_rules = {}

    # 更新後のルールを適用
    for key, rules in updated_rules.items():
        sheet.conditional_formatting._cf_rules[key] = rules

"""数式に含まれる行番号を+1する（影響を受ける行のみ）"""
def increment_formula_row_numbers(formula, insert_row):
    """
    - `$A9` → `$A10`
    - `$A$9` → `$A$10`
    - `A9` → `A10`
    """
    def replace_match(match):
        col, dollar_sign, row = match.groups()
        row_num = int(row)
        # 挿入行より下の行のみ+1
        if row_num >= insert_row:
            return f"{col}{dollar_sign}{row_num + 1}"
        return match.group(0)  # 変更なし(パターンにマッチした文字列全体を返す)

    # A1, $A1, $A$1 の形式を検出
    return re.sub(r"(\$?[A-Z]+)(\$?)(\d+)", replace_match, formula)

"""勤務イベントを日ごとに整理し、勤務時間を計算"""
def process_work_events(events):
    attendance = {}
    for event in events:
        try:
            start = event['start'].get('dateTime', event['start'].get('date'))
            end = event['end'].get('dateTime', event['end'].get('date'))
            start_dt = datetime.datetime.fromisoformat(start)
            end_dt = datetime.datetime.fromisoformat(end)
            location = event.get('location', '')
            description = event.get('description', '')
            title = event.get('summary', '')
            date_key = start_dt.date()
            # 取得したイベントの中から終日イベントを判定して出力する
            if is_all_day_event(event):
                allday = '終日'
            else:
                allday = ''
            # # 休憩イベントの場合（例：「休憩」や「ランチ」など）
            # if "休憩" in event.get("summary", "") or "ランチ" in event.get("summary", ""):
            #     attendance[date_key]["break"] += (end_dt - start_dt).seconds / 3600  # 時間換算
            # else:
            #     # 出勤時間の記録
            #     if attendance[date_key]["start"] is None or start_dt < attendance[date_key]["start"]:
            #         attendance[date_key]["start"] = start_dt
            #     # 退勤時間の記録
            #     if attendance[date_key]["end"] is None or end_dt > attendance[date_key]["end"]:
            #         attendance[date_key]["end"] = end_dt
            if date_key not in attendance:
                attendance[date_key] = {'start': start_dt,
                                        'end': end_dt,
                                        # 'total_hours': (end_dt - start_dt).seconds / 3600,
                                        # 'location': location,
                                        # 'description': description,
                                        # 'title': title,
                                        # 'allday': allday
                                        }
            else:
                if allday == '終日':
                    attendance[date_key]['start'] = start_dt
                    attendance[date_key]['end'] = end_dt
                elif not attendance[date_key].get('allday'):
                    # 既存の出退勤と比較して、出勤は最も早い時刻、退勤は最も遅い時刻をセット
                    attendance[date_key]['start'] = min(attendance[date_key]['start'], start_dt)
                    attendance[date_key]['end'] = max(attendance[date_key]['end'], end_dt)
                # attendance[date_key]['total_hours'] += (end_dt - start_dt).seconds / 3600  # 勤務時間の合計
            if not attendance[date_key].get('location'):
                attendance[date_key]['location'] = location
            if not attendance[date_key].get('description'):
                attendance[date_key]['description'] = description
            if not attendance[date_key].get('title'):
                attendance[date_key]['title'] = title
            if allday == '終日':
                attendance[date_key]['allday'] = allday
        except Exception:
            logger.exception('Exception')
    # 合計勤務時間を計算
    for date, record in attendance.items(): # 各要素のキーkeyと値value
        try:
            if record["start"] and record["end"]:
                total_hours = (record["end"] - record["start"]).seconds / 3600
                # total_hours = (record["end"] - record["start"]).seconds / 3600 - record["break"]
                attendance[date]["total_hours"] = round(total_hours, 2)
        except Exception:
            logger.exception('Exception')

    return attendance

"""勤務時間を個別に集計"""
def create_attendance_data(events):
    attendance = {}

    for event in events:
        try:
            start = event['start'].get('dateTime', event['start'].get('date'))
            end = event['end'].get('dateTime', event['end'].get('date'))

            start_dt = datetime.datetime.fromisoformat(start)
            end_dt = datetime.datetime.fromisoformat(end)
            date_key = start_dt.date()
            location = event.get('location', '')
            description = event.get('description', '')
            title = event.get('summary', '')
            # 取得したイベントの中から終日イベントを判定して出力する
            if is_all_day_event(event):
                allday = '終日'
            else:
                allday = ''
            if date_key not in attendance:
                attendance[date_key] = []
            attendance[date_key].append({'start': start_dt,
                                        'end': end_dt,
                                        'location': location,
                                        'description': description,
                                        'title': title,
                                        'allday': allday
                                        })
        except Exception:
            logger.exception('Exception')

    return attendance
# イベントが終日かどうかを判定
def is_all_day_event(event):
    if event['start']:
        return 'date' in event['start']  # 'date' があれば終日イベント
    else:
        return False
# 例: Google Calendar API から取得したイベント
# event_all_day = {
#     'start': {'date': '2025-02-06'},
#     'end': {'date': '2025-02-07'}
# }
# event_timed = {
#     'start': {'dateTime': '2025-02-06T09:00:00+09:00'},
#     'end': {'dateTime': '2025-02-06T18:00:00+09:00'}
# }
